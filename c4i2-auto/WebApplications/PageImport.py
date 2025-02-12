import os
import time

import pandas as pd
import pyautogui
import pyperclip
from selenium.webdriver.common.by import By

from Libraries.Framework.Paths import Paths
from Libraries.Framework.Utils import PageBase
from Libraries.Plugins.PandasExcel import PandasExcel
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from Libraries.Config import Import


class ImportPageObject(PageBase):
    """
    Class chứa các hàm thực hiện tương tác với web để verify value import
    """

    def __init__(self, driver):
        super().__init__(driver)
        self.pathDatas = Paths().get_path_datas()

    def enter_data_for_search(self, item1, item2):
        """
        :param item1: header for search
        :param item2: value for search
        :return: Hành động nhập thông tin record muốn tìm vào tab Tìm kiếm
        """
        button_search = f'((//div[text()="{item1}"])//ancestor::div[contains(@class, "form-control-label")])//input'
        self.send_key_input(button_search, item2)
        time.sleep(0.5)

    def click_button_search(self):
        """
        :return: Hành động nhấn vào button "Tìm kiếm"
        """
        time.sleep(1)
        label = "Tìm kiếm"
        btn_tim_kiem = f'//button[.//span[text()="{label}"]]'
        self.click_obj(btn_tim_kiem)
        time.sleep(2)

    def check_edit_icon(self):
        """
        Kiểm tra icon edit trên grid
        :return: not_find or find_1 or find_n
        """
        icon_edit = '(//i[contains(@class,"fal fa-edit")])'
        icon_edit_elements = self.driver.find_elements(By.XPATH, icon_edit)
        icon_edit_count = len(icon_edit_elements)
        icon_edit = '(//i[contains(@class,"fal fa-edit")])[1]'
        if icon_edit_count == 0:
            return "not_found"
        elif icon_edit_count == 1:
            self.click_obj(icon_edit)
            time.sleep(0.5)
            return "found_single"
        else:
            return "found_multiple"

    def click_close_popup(self):
        """
        :return: Hành động nhấn đóng popup sau khi muốn view
        """
        time.sleep(0.5)
        icon_close_pop_up = '//button[contains(@class,"btn--xs btn--only-icon")]'
        btn_xac_nhan = '//span[contains(text(), "Xác nhận")]'
        btn_close_not_verify = '//button[contains(@class,"btn--default btn--round")]'
        try:
            self.click_obj(icon_close_pop_up)
            time.sleep(0.5)
            self.click_obj(btn_xac_nhan)
            time.sleep(0.5)
        except Exception as e:
            self.click_obj(btn_close_not_verify)
            time.sleep(0.5)
            print("Đã xảy ra lỗi:", str(e))

    def do_action_delete(self):
        """
        :return: True(Xóa thành công), False(Không tìm thấy record muốn xóa)
        """
        checkbox_delete = '(//div[contains(@class, "data-grid-wrapper")]//child::span[@class="checkbox-input checkbox--icon-md"])[2]'
        if self.check_element_visibility(checkbox_delete):
            btn_delete = '//button[contains(@class, "btn--danger")]'
            btn_xac_nhan = '//div[@class= "popup-footer"]//child::button[contains(@class, "btn btn--primary")]'
            self.click_obj(checkbox_delete)
            time.sleep(0.5)
            self.click_obj(btn_delete)
            time.sleep(0.5)
            self.click_obj(btn_xac_nhan)
            time.sleep(0.5)
        else:
            return False


class PageImport(PageBase):
    def __init__(self, driver):
        super().__init__(driver)
        self.pathDatas = Paths().get_path_datas()
        self.pathReports = Paths().get_path_reports()

    def private_fill_file_import(self, type_of, label, val):
        if type_of == "dropdown-btn-text":
            btn_arrow = f'//div[contains(@class,"popup")]//div[contains(@class,"form-control-label") and .//div[text()="{label}"]]//i[contains(@class,"far fa-chevron")]'
            # self.click_obj(btn_arrow)
            # time.sleep(1)
            # txt_val_select = f'//div[@class="as-dropdown-item-button" and text()="{val}"]'
            # self.click_obj(txt_val_select)
            self.select_key_dropdown(btn_arrow, val)

            btn_huy = '//span[text()="Hủy"]'
            self.click_obj(btn_huy)
            time.sleep(2)
        if type_of == "file":
            path_excel = os.path.join(self.pathDatas, "c4i2_Data_Import", label)
            self.driver.find_element(By.XPATH, '//input[@type="file" and @id="upload-excel"]').send_keys(path_excel)
            time.sleep(2)
        if type_of == "folder":
            time.sleep(1)
            path_folder = os.path.join(self.pathDatas, "c4i2_Data_Import", label)
            pyperclip.copy(path_folder)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            # Nhấn Enter lần 1
            pyautogui.press('enter')

            # Nhấn Enter lần 2
            pyautogui.press('enter')
            time.sleep(0.5)

            pyautogui.press('tab')
            pyautogui.press('enter')

    def do_verify_import_status(self, filename, status):
        time.sleep(3)
        # verify status
        check_status_done = (f'//div[contains(@class, "list-item") and .//div[text()="{filename}"]]//child::span[text()="{status}"]')
        if self.check_element_visibility(check_status_done):
            return True
        else:
            return False

    def do_verify_import_data_status_for_row(self, status_import, value_import):
        time.sleep(3)
        # verify status
        value_import = value_import[0]
        check_status = f'//div[contains(text(), "{value_import}")]//ancestor::div[contains(@class, "dg-row")]//child::span[text()="{status_import}"]'
        if self.check_element_visibility(check_status):
            return True
        else:
            return False

    def do_verify_import_in_popup(self, type_of, label, val):
        if type_of == "file":
            import_page_object = ImportPageObject(self.driver)
            filename = label
            sheetname = val
            path_file = os.path.join(self.pathDatas, "c4i2_Data_Import", filename)
            pe = PandasExcel(path_file, sheetname)
            df = pe.get_data_excel()

            result_in_excel, required_data = self.process_file_excel(filename, sheetname)
            number_value = result_in_excel[0].__len__()

            # Danh sách lưu trữ các giá trị được import thành công - hỗ trợ xóa dữ liệu sau khi verify
            list_import_success = []

            # Tải workbook từ file Excel đã tồn tại
            workbook = load_workbook(path_file)
            worksheet = workbook[sheetname]

            print(f"Số lượng thuộc tính import {number_value}")
            for i, item in enumerate(required_data):
                print(f"Bắt đầu kiểm tra dữ liệu import dòng thứ {i + 1}")
                for header, value in item:
                    item1 = str(header)
                    item2 = str(value)
                    import_page_object.enter_data_for_search(item1, item2)

                import_page_object.click_button_search()

                status_icon_edit = import_page_object.check_edit_icon()

                if status_icon_edit == "found_single":
                    result_in_popup = self.get_value_in_pop_up()

                    # Lấy giá trị từng hàng trong file excel ra để so sánh
                    value_in_excel = result_in_excel[i]

                    found_match = False  # Biến hỗ trợ lấy value import success

                    # 2 mảng hỗ trợ in kết quả so sánh ra màn hình
                    mang_imported = 0
                    mang_unimported = 0

                    # Bắt đầu so sánh kết quả file excel và giá trị trong popup
                    for item in value_in_excel:
                        header = item[0]  # Tiêu đề
                        value = str(item[1])  # Giá trị

                        if header == "File ảnh khuôn mặt":
                            name_image_import = value
                            image_xpath = '//div[@class="upload-image-container"]//img'
                            # Gọi hàm so sánh ảnh
                            check = self.compare_images_import(name_image_import, image_xpath)
                            if check:
                                mang_imported += 1
                                value_success = True
                            else:
                                mang_unimported += 1
                                value_success = False
                        else:
                            if (header, value) in result_in_popup:
                                value_success = True
                                mang_imported += 1
                                if not found_match:
                                    found_match = True
                                    list_import_success.append(required_data[i])
                            else:
                                value_success = False
                                mang_unimported += 1
                                # Tô màu hồng cho ô nếu value_success = False
                        if not value_success:
                            for column_name in df.columns:
                                if column_name.startswith(header):
                                    column_index = df.columns.get_loc(column_name)
                                    break
                            row_index = i + 2  # Vì hàng đầu tiên là header nên cộng 2
                            cell = worksheet.cell(row=row_index, column=column_index + 1)
                            cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                    fill_type="solid")  # Màu hồng

                    if mang_unimported == 0:
                        import_success = "Imported"
                    elif mang_unimported == number_value:
                        import_success = "Not Imported"
                    else:
                        import_success = "Incomplete Import"

                    # importPageObject.clickClosePopup()
                    self.driver.refresh()
                    time.sleep(4)
                    print(f"- Số thuộc tính đã được Import {mang_imported}/{number_value}")
                    print(f"- Số thuộc tính chưa được Import {mang_unimported}/{number_value}")

                elif status_icon_edit == "not_found":
                    print(f'Dữ liệu dòng thứ {i + 1} không được import')
                    import_success = "Not Imported"
                else:
                    print(f"Dòng thứ {i + 1} dữ liệu import trùng")
                    # Biến đánh dấu trạng thái import cho record
                    import_success = "Not Imported"

                self.update_and_color_import_status(df, worksheet, import_success)

            # Tạo file excel -> Kết quả Import
            path_file_new = os.path.join(self.pathReports, "Web", "ResultImport")
            file_path = os.path.join(path_file_new, "result.xlsx")
            # Lưu file Excel đã được chỉnh sửa
            workbook.save(file_path)

            # Trả về kết quả trạng thái import, danh sách import thành công
            return list_import_success, file_path

    @staticmethod
    def update_and_color_import_status(df, worksheet, import_success):
        """
        :param df: DataFrame đang thao tác
        :param worksheet: Bảng tính đang tạo mới
        :param import_success: Trạng thái record đã import (True/False)
        :return: Giá trị Imported và Not Imported ở worksheet và màu sắc tương ứng
        """
        columnstatus = Import.NameColumnStatus
        if columnstatus not in df.columns:
            df[columnstatus] = ""  # Tạo cột mới với giá trị mặc định là rỗng
        # Đảm bảo có đủ cột trong worksheet để chứa cột mới
        if worksheet.max_column < len(df.columns):
            worksheet.insert_cols(worksheet.max_column + 1, amount=len(df.columns) - worksheet.max_column)

        # Đặt tên cho cột mới là "Status Import"
        worksheet.cell(row=1, column=len(df.columns), value=columnstatus)
        # Lặp qua từng hàng trong DataFrame
        for i, row in df.iterrows():
            # Kiểm tra điều kiện để xác định trạng thái "import_success"
            if import_success == "Imported":
                import_status = Import.Imported
                cell_color = "00FF00"  # Màu xanh lá
            elif import_success == "Incomplete Import":
                import_status = Import.IncompleteImport
                cell_color = "FFFF00"  # Màu vàng
            else:
                import_status = Import.NotImported
                cell_color = "FF0000"  # Màu đỏ
            worksheet.cell(row=i + 2, column=df.columns.get_loc(columnstatus) + 1, value=import_status)
            worksheet.cell(row=i + 2, column=df.columns.get_loc(columnstatus) + 1).fill = PatternFill(
                start_color=cell_color, end_color=cell_color, fill_type="solid")

    def compare_images_import(self, name_image_import, image_xpath):
        """
        :param name_image_import: Tên hình ảnh đã import
        :param image_xpath: Xpath của hình ảnh trên web
        :return: True (match image), False (not match image)
        """
        # Lấy thuộc tính 'src' của phần tử hình ảnh
        image_element = self.driver.find_element(By.XPATH, image_xpath)
        image_src = image_element.get_attribute('src')

        # Tìm giá trị sau '?' và trước '&'
        match = re.search(r'\?(.*?)&', image_src)
        if not match:
            return False

        value = match.group(1)

        # Kiểm tra nếu giá trị chứa "fileName="
        if "fileName=" in value:
            value = value.split("fileName=")[-1]
        else:
            name_image_import = "id=undefined"

        if value == name_image_import:
            return True

        return False

    def process_file_excel(self, filename, sheetname):
        """
        :param filename: Tên file excel đã import
        :param sheetname: Tên sheetname chứa dữ liệu import
        :return: 2 mảng (header-value): All & Required
        """
        path_file = os.path.join(self.pathDatas, "c4i2_Data_Import", filename)
        pe = PandasExcel(path_file, sheetname)
        df = pe.get_data_excel()

        # Loại bỏ các cột chứa giá trị NaN
        # df = df.dropna(axis=1)

        # Lấy danh sách tiêu đề từ DataFrame
        titles_excel = df.columns.tolist()

        result_in_excel = []
        for index, row in df.iterrows():
            pairs = []
            for header in titles_excel:
                header_modified = header.replace('*', '')
                value = row[header]
                if pd.notna(value):
                    pair = (header_modified, value)
                else:
                    pair = (header_modified, "")
                pairs.append(pair)
            result_in_excel.append(pairs)

        required_column = []
        for columnname in titles_excel:
            if '*' in columnname:
                modified_column = columnname.replace('*', '')
                required_column.append(modified_column)

        # Tạo mảng lưu header và value của Required Column trong ResultInExcel
        required_data = []
        for row in result_in_excel:
            data = []
            for header, value in row:
                if header in required_column:
                    pair = (header, value)
                    data.append(pair)
            required_data.append(data)

        return result_in_excel, required_data

    def get_value_in_pop_up(self):
        """
        :return: Lấy giá trị (header-value) trong popup
        """
        # Xpath kết quả popup chi tiếc
        div_elements = self.driver.find_elements(By.XPATH,
                                                 '//div[contains(@class, "popup")]//div[contains(@class, "form-control-label")]')
        result_in_popup = []
        for div in div_elements:
            html_string = div.get_attribute("outerHTML")
            soup = BeautifulSoup(html_string, 'html.parser')
            texts = [text.strip() for text in soup.stripped_strings]
            if len(texts) == 1:
                if "<input" in html_string:
                    input_element = div.find_element(By.XPATH, './/input')
                    value = input_element.get_attribute("value")
                    text = texts[0]
                    if text and value:
                        result_in_popup.append((text, value))
            else:
                text = texts[0]
                value = texts[1]
                result_in_popup.append((text, value))

        return result_in_popup

    def do_delete_data_import(self, list_import_success):
        """
        :param list_import_success: Danh sách giá trị (heade-value) đã import thành công
        :return: Hành động search và xóa các giá trị đã import
        """
        import_page_object = ImportPageObject(self.driver)
        for i, item in enumerate(list_import_success):
            for header, value in item:
                item1 = header  # Tiêu đề để search
                item2 = value  # Giá trị để search
                button_search = f'((//div[text()="{item1}"])//ancestor::div[contains(@class, "form-control-label")])//input'
                self.send_key_input(button_search, item2)
                time.sleep(0.5)

            time.sleep(1)
            label = "Tìm kiếm"
            btn_tim_kiem = f'//button[.//span[text()="{label}"]]'
            self.click_obj(btn_tim_kiem)
            time.sleep(2)
            # Gọi hàm thực hiện hành động xóa
            import_page_object.do_action_delete()

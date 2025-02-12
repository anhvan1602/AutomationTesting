import os.path
import os
import platform
import subprocess
import webbrowser
from datetime import datetime
import openpyxl
import pandas as pd
from Libraries.Framework.Paths import Paths


class PandasExcel:
    def __init__(self, path_data_file, sheet_name):
        self.pathDataFile = path_data_file
        self.sheet_name = sheet_name

    def insert_new_value(self, label, new_value):
        wb = openpyxl.load_workbook(self.pathDataFile)
        sheet = wb[self.sheet_name]

        found_label = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == label:
                    found_label = True
                    col_index = cell.column
                    row_index = cell.row + 1
                    max_row = sheet.max_row
                    for i in range(max_row, row_index - 1, -1):
                        sheet.cell(row=i + 1, column=col_index).value = sheet.cell(row=i, column=col_index).value
                    sheet.cell(row=row_index, column=col_index).value = new_value
                    wb.save(self.pathDataFile)

                    return

        if not found_label:
            print(f"Label '{label}' not found in the worksheet.")

    def get_data_excel(self):
        return pd.read_excel(self.pathDataFile, sheet_name=self.sheet_name)

    @staticmethod
    def get_row_data_file(df):
        return len(df.axes[0])

    def get_row_by_index(self, index_row_path):
        result = dict()
        df = self.get_data_excel()
        rows = self.get_row_data_file(df)

        if index_row_path > rows - 1:
            print('\nTong so dong dang co:  ', rows, ' | So dong can lay ra:  ', index_row_path + 1)
        else:
            result["TypeOf"] = df['TypeOf'][index_row_path]
            result["Label"] = df['Label'][index_row_path]
            result["Value"] = df['Value'][index_row_path]
        return result

    def get_val_by_index(self, index_row, col_name):
        return self.get_row_by_index(index_row)[col_name]

    def get_val_by_index1(self, index_row, col_name):
        """
        Lấy giá trị trong file Excel dựa trên chỉ số dòng và tên cột
        :param index_row: chỉ số dòng (bắt đầu từ 1)
        :param col_name: tên cột
        :return: giá trị ô tương ứng
        """
        workbook = openpyxl.load_workbook(self.pathDataFile)
        sheet = workbook[self.sheet_name]
        col_index = None
        for cell in sheet[1]:
            if cell.value == col_name:
                col_index = cell.column
                break
        if col_index is None:
            raise ValueError(f"Không tìm thấy cột {col_name}")
        for row in sheet.iter_rows(min_row=index_row, max_row=index_row):
            for cell in row:
                if cell.column == col_index:
                    return cell.value
        raise ValueError(f"Không tìm thấy dòng {index_row}")


def save_to_excel(lst_data, file_name, open_file=True):
    if len(lst_data) > 0:
        df = pd.DataFrame(lst_data, columns=lst_data[0].keys())
        path_file = os.path.join(Paths().get_path_reports(), f'{_get_string_now()}_{file_name}')
        df.to_excel(path_file, index=False)
        print("Đã lưu file Excel:", path_file)

        if open_file:
            try:
                _open_excel(path_file)
            except Exception as e:
                print("Không thể mở file Excel:", path_file)
                print("Lỗi:", str(e))
    else:
        print("Dữ liệu rỗng. Không có gì để lưu.")


def _get_string_now():
    now = datetime.now()
    return now.strftime("%d%m%y_%H%M%S")


def _open_excel(path_file):
    system = platform.system()
    if system == 'Windows':
        os.startfile(path_file)
    elif system == 'Darwin':
        subprocess.call(['open', path_file])
    else:
        webbrowser.open(path_file)


if __name__ == "__main__":
    # fileName = "C4i2_Datas_v1.0.xlsx"
    # sheetName_ThemPhuongTien = "VuViec_ThemPhuongTien"
    #
    # pathsDatas = Paths().getPathDatas()
    # pathFile = os.path.join(pathsDatas, fileName)
    #
    # pe = PandasExcel(pathFile, sheetName_ThemPhuongTien)
    # df = pe.getDataExcel()
    #
    # i = 0
    # while i < int(pe.getRowDataFile(df)):
    #     typeOf = df['TypeOf'][i]
    #     label = df['Label'][i]
    #     lstAttribute = df['Value'][i]
    #     print(typeOf)
    #     print(label)
    #     print(lstAttribute)
    #     i = i + 1
    # Ví dụ danh sách đối tượng
    data = [
        {"Name": "John", "Age": 25, "Country": "USA"},
        {"Name": "Alice", "Age": 30, "Country": "Canada"},
        {"Name": "Bob", "Age": 35, "Country": "UK"}
    ]

    filename = "output.xlsx"  # Tên file Excel đầu ra
    save_to_excel(data, filename)
